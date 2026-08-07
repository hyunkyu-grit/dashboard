"""
Load market data directly from the Infomax Excel export workbook (e.g. Total Data.xlsx).

Sheet map:
  CD91 4사 평균 AAA                  -> CD 91D rate, value column index 1
  IRS KRW (Qr VS Qr 91D CD91) 0XY  -> IRS par rate, MID column index 3 (bid/ask/mid)

Layout: long format (dates as rows). Row 0-2 are header/metadata, data starts at
row index 3, descending date order, padded with tens of thousands of empty rows
beyond the real data -- iteration stops at the first blank date cell per sheet.

Rates are stored as percentage units (e.g. 2.92 = 2.92%); normalized to decimal.
"""

from __future__ import annotations

import os
import shutil
import tempfile
import time
from datetime import date, datetime
from pathlib import Path

from ..core.errors import _check_business_day
from ..core.market_data import MarketSnapshot, RateQuote
from .cache import get_cached

_CD_SHEET = "CD91 4사 평균 AAA"
_CD_COL = 1

_IRS_SHEETS: dict[str, int] = {
    "IRS KRW (Qr VS Qr 91D CD91) 01년": 1,
    "IRS KRW (Qr VS Qr 91D CD91) 02년": 2,
    "IRS KRW (Qr VS Qr 91D CD91) 03년": 3,
    "IRS KRW (Qr VS Qr 91D CD91) 05년": 5,
    "IRS KRW (Qr VS Qr 91D CD91) 07년": 7,
    "IRS KRW (Qr VS Qr 91D CD91) 10년": 10,
}
_IRS_COL = 3  # MID 종가

_HEADER_ROWS = 3


def _parse_cell_date(value) -> date | None:
    """Normalize a date cell to a date object: native datetime, int YYYYMMDD,
    or string (YYYY-MM-DD[ HH:MM:SS] / YYYYMMDD)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, int):
        s = str(value)
        if len(s) == 8:
            try:
                return datetime.strptime(s, "%Y%m%d").date()
            except ValueError:
                return None
        return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None
    return None


def _normalize_rate(value: float) -> float:
    """Detect percentage (e.g. 2.92) vs decimal (e.g. 0.0292) and return decimal."""
    return value / 100.0 if abs(value) > 1 else value


def _open_workbook(excel_path: Path):
    """Open the workbook read-only. If the file is locked (Infomax keeps it open),
    copy it to a temp file first and read the copy instead.

    Returns (workbook, tmp_path); tmp_path is None unless a temp copy was made,
    in which case the caller must delete it after closing the workbook.
    """
    if excel_path.suffix.lower() == ".xls":
        return _open_xls(excel_path)

    import openpyxl

    try:
        return openpyxl.load_workbook(excel_path, read_only=True, data_only=True), None
    except PermissionError:
        tmp_path = Path(tempfile.gettempdir()) / f"_irs_pricer_xlcopy_{os.getpid()}_{excel_path.name}"
        last_err: Exception | None = None
        for attempt in range(3):
            try:
                shutil.copy2(excel_path, tmp_path)
                last_err = None
                break
            except PermissionError as e:
                last_err = e
                time.sleep(0.3 * (attempt + 1))
        if last_err is not None:
            raise PermissionError(
                f"{excel_path} is locked and could not be copied for reading "
                f"(file may be held with an exclusive, read-denying lock)"
            ) from last_err
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        return wb, tmp_path


def _open_xls(excel_path: Path):
    """Legacy .xls fallback via xlrd, wrapped to expose the same sheet-by-name,
    iter_rows(values_only=True) interface used for .xlsx below."""
    import xlrd

    class _XlrdSheetAdapter:
        def __init__(self, sheet):
            self._sheet = sheet

        def iter_rows(self, values_only: bool = True):
            for r in range(self._sheet.nrows):
                yield tuple(self._sheet.cell_value(r, c) for c in range(self._sheet.ncols))

    class _XlrdWorkbookAdapter:
        def __init__(self, book):
            self._book = book

        def __getitem__(self, name: str):
            return _XlrdSheetAdapter(self._book.sheet_by_name(name))

        def close(self):
            pass

    try:
        book = xlrd.open_workbook(str(excel_path))
        return _XlrdWorkbookAdapter(book), None
    except PermissionError:
        tmp_path = Path(tempfile.gettempdir()) / f"_irs_pricer_xlcopy_{os.getpid()}_{excel_path.name}"
        shutil.copy2(excel_path, tmp_path)
        book = xlrd.open_workbook(str(tmp_path))
        return _XlrdWorkbookAdapter(book), tmp_path


def _close_workbook(wb, tmp_path: Path | None) -> None:
    wb.close()
    if tmp_path is not None:
        try:
            tmp_path.unlink()
        except OSError:
            pass


def _read_sheet_series(ws, value_col: int) -> dict[date, float]:
    """Read a {date: value} series from a long-format Infomax sheet, stopping
    at the first blank date cell after the header (avoids scanning padding rows)."""
    series: dict[date, float] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < _HEADER_ROWS:
            continue
        if row[0] is None:
            break
        d = _parse_cell_date(row[0])
        if d is None:
            continue
        val = row[value_col]
        if val is None:
            continue
        series[d] = float(val)
    return series


def _parse_all_series(excel_path: Path) -> dict[str, dict[date, float]]:
    """Parse every sheet this module needs (CD91 + each IRS tenor sheet) in
    one workbook open, returning {sheet_name: {date: raw_value}}. This is
    the function that actually pays the openpyxl parse cost; callers go
    through _load_all_series() below so repeated lookups -- any date, any
    tenor, common_dates_xl, load_fixing_history_xl -- share a single cached
    parse instead of each re-opening and re-scanning the workbook."""
    wb, tmp_path = _open_workbook(excel_path)
    try:
        series = {_CD_SHEET: _read_sheet_series(wb[_CD_SHEET], _CD_COL)}
        for sheet_name in _IRS_SHEETS:
            series[sheet_name] = _read_sheet_series(wb[sheet_name], _IRS_COL)
        return series
    finally:
        _close_workbook(wb, tmp_path)


def _load_all_series(excel_path: Path) -> dict[str, dict[date, float]]:
    return get_cached(Path(excel_path), _parse_all_series)


def common_dates_xl(excel_path: Path) -> list[date]:
    """Return all dates present in the CD91D sheet AND every IRS tenor sheet, sorted ascending."""
    excel_path = Path(excel_path)
    all_series = _load_all_series(excel_path)
    common = set(all_series[_CD_SHEET].keys())
    for sheet_name in _IRS_SHEETS:
        common &= set(all_series[sheet_name].keys())
    if not common:
        raise ValueError(f"{excel_path.name}의 모든 시트에 공통된 날짜가 없습니다.")
    return sorted(common)


def latest_common_date_xl(excel_path: Path) -> date:
    """Return the most recent date present in the CD91D sheet AND every IRS tenor sheet."""
    return common_dates_xl(excel_path)[-1]


def load_market_snapshot_xl(excel_path: Path, valuation_date: date) -> MarketSnapshot:
    """Build a MarketSnapshot from the Infomax Excel export for valuation_date.

    Raises NonBusinessDayError if the date is a weekend or public holiday.
    Raises ValueError if data is missing for a valid business day.
    """
    _check_business_day(valuation_date)

    excel_path = Path(excel_path)
    all_series = _load_all_series(excel_path)

    raw_cd = all_series[_CD_SHEET].get(valuation_date)
    if raw_cd is None:
        raise ValueError(f"{excel_path.name}에 {valuation_date}의 CD91D 금리가 없습니다.")
    cd_rate = _normalize_rate(raw_cd)

    swap_quotes: list[RateQuote] = []
    for sheet_name, tenor in _IRS_SHEETS.items():
        raw_irs = all_series[sheet_name].get(valuation_date)
        if raw_irs is None:
            raise ValueError(f"{excel_path.name}에 {valuation_date}의 IRS {tenor}Y MID 금리가 없습니다.")
        swap_quotes.append(RateQuote(tenor_years=tenor, rate=_normalize_rate(raw_irs)))

    return MarketSnapshot(
        valuation_date=valuation_date,
        cd_rate=cd_rate,
        swap_quotes=swap_quotes,
    )


def load_fixing_history_xl(excel_path: Path) -> dict[date, float]:
    """Return {date: rate(decimal)} for the CD91D sheet, usable as
    `historical_fixings` in mtm_valuation."""
    excel_path = Path(excel_path)
    raw = _load_all_series(excel_path)[_CD_SHEET]
    return {d: _normalize_rate(v) for d, v in raw.items()}


_RESULT_SHEET = "IRS_Result"
_RESULT_COLUMNS = [
    "valuation_date",
    "tenor",
    "notional",
    "fixed_rate",
    "pay_fixed",
    "trade_date",
    "remaining_tenor",
    "par_rate",
    "npv",
    "fixed_leg_pv",
    "float_leg_pv",
    "dv01",
    "mtm_pnl",
    "timestamp",
]

_RATE_COLS = {"fixed_rate", "par_rate"}
_KRW_COLS = {"notional", "npv", "fixed_leg_pv", "float_leg_pv", "dv01", "mtm_pnl"}
_DATE_COLS = {"valuation_date", "trade_date"}


def export_result_xl(export_path: Path, row: dict) -> None:
    """Append one pricing-result row to `export_path` (sheet "IRS_Result"),
    creating the file and header row if it doesn't already exist."""
    import openpyxl

    export_path = Path(export_path)
    if export_path.exists():
        wb = openpyxl.load_workbook(export_path)
        if _RESULT_SHEET in wb.sheetnames:
            ws = wb[_RESULT_SHEET]
        else:
            ws = wb.create_sheet(_RESULT_SHEET)
            ws.append(_RESULT_COLUMNS)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _RESULT_SHEET
        ws.append(_RESULT_COLUMNS)

    out_row = []
    for col in _RESULT_COLUMNS:
        out_row.append(row.get(col))
    ws.append(out_row)

    next_row = ws.max_row
    for col_idx, col in enumerate(_RESULT_COLUMNS, start=1):
        cell = ws.cell(row=next_row, column=col_idx)
        if col in _RATE_COLS:
            cell.number_format = "0.0000%"
        elif col in _KRW_COLS:
            cell.number_format = "#,##0"
        elif col in _DATE_COLS:
            cell.number_format = "yyyy-mm-dd"

    wb.save(export_path)
    wb.close()
