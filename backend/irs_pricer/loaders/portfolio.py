"""
Load booked-position data from "Portfolio Data.xlsx" -- a combined bond +
IRS workbook. Ported from rates-simulator-main's ExcelUploader.tsx, the one
already-working position-Excel parser in either codebase
(components/ExcelUploader.tsx:164-376): a bond sheet (name containing
"채권", else the first sheet) plus an IRS sheet (name containing "IRS").

Bond rows keep their own pre-computed fields from the source ledger
(evaluation amount, market yield, duration, a duration-weighted PVBP) --
this backend has no bond pricing engine, so bond figures are read straight
from the file rather than computed. IRS rows are booked-swap-shaped
(position_id, start_date, maturity_date, notional, fixed_rate, pay_fixed,
float_spread) for the existing pricing engine; their PVBP is computed
separately by portfolio_analytics_service via the real delta engine, not
here.

Both row shapes carry `instrument_type` ("bond"|"irs"), `sector`, and
`book` so the caller can group/aggregate across the whole mixed portfolio.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

XLSX_NAME = "Portfolio Data.xlsx"

_NAME_START_DATE_RE = re.compile(r"(\d{6})-(\d{6})")

# Priority order mirrors the source (:309): first non-empty wins.
_FIXED_RATE_COLUMNS = ["구조화쿠폰", "계약이율", "표면이율", "고정금리", "이율"]

# 상품소분류명 -> the task's 7-sector taxonomy (ExcelUploader.tsx:179-190).
_BOND_SECTOR_MAP = {
    "국고채": "국고채",
    "통안채": "통안채",
    "특수은행채": "특은채",
    "일반은행채": "시은채",
    "공사공단채": "공사채",
    "비금융특수채": "공사채",
    "지방공사특수채": "공사채",
    "금융회사채": "여전채",
    "일반사채": "회사채",
}


def _to_date(value: object) -> date | None:
    """Excel date cells already come back as datetime/date from openpyxl;
    unlike the TS source (which parses raw serials via parseExcelDate), no
    manual Excel-serial parsing is needed here."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _to_text_date(value: object) -> date | None:
    """The bond sheet's 발행일자/만기일자 arrive as apostrophe-prefixed text
    (e.g. "'2026-11-22" -- the ledger export forces text format, and openpyxl
    hands the apostrophe back as part of the value). Strip it and parse
    strictly: anything that isn't a real ISO date becomes None, never a
    guessed date -- a wrong date prices wrong, a None only degrades to the
    analytic fallback. Real date cells pass through, in case a future export
    drops the text formatting."""
    d = _to_date(value)
    if d is not None:
        return d
    if not isinstance(value, str):
        return None
    text = value.strip().lstrip("'").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_name_start_date(name: str) -> date | None:
    """Port of the source's parseNameStartDate (:30-39): a YYMMDD-YYMMDD
    token embedded in the position name, first group is the start date."""
    m = _NAME_START_DATE_RE.search(name or "")
    if not m:
        return None
    s = m.group(1)
    yy, mm, dd = int(s[0:2]), int(s[2:4]), int(s[4:6])
    if not (1 <= mm <= 12) or not (1 <= dd <= 31):
        return None
    try:
        return date(2000 + yy, mm, dd)
    except ValueError:
        return None


def _map_book(raw: object) -> str:
    """RP Fund is both the bond sheet's own default (:230, blank -> 'RP
    Fund') and the IRS sheet's mapped value for any '파생'-containing book
    name (:287) -- unified into one rule here since both mean the same desk
    book in this ledger, and a blank IRS book name is treated the same way
    for consistency (the source leaves it as '', which isn't a usable
    grouping key)."""
    raw_book = str(raw or "").strip()
    if not raw_book or "파생" in raw_book:
        return "RP Fund"
    return raw_book


def _map_bond_sector(sub_class: object) -> str:
    return _BOND_SECTOR_MAP.get(str(sub_class or "").strip(), "기타")


def _bucket_for_years(years: float) -> str:
    """Nearest-tenor bucket for a bond's remaining maturity, extending the
    source's thresholds (ExcelUploader.tsx:192-203, which skips 6Y/8Y/9Y)
    to cover the full 16-column tenor grid this task's PVBP table needs."""
    if years <= 1 / 365:
        return "1D"
    if years <= 0.25:
        return "3M"
    if years <= 0.5:
        return "6M"
    if years <= 0.75:
        return "9M"
    if years <= 1:
        return "1Y"
    if years <= 1.5:
        return "1.5Y"
    if years <= 2:
        return "2Y"
    if years <= 3:
        return "3Y"
    if years <= 4:
        return "4Y"
    if years <= 5:
        return "5Y"
    if years <= 6:
        return "6Y"
    if years <= 7:
        return "7Y"
    if years <= 8:
        return "8Y"
    if years <= 9:
        return "9Y"
    if years <= 10:
        return "10Y"
    return "30Y"


def _find_irs_sheet(wb) -> str | None:
    for name in wb.sheetnames:
        if "IRS" in name:
            return name
    return None


def _find_bond_sheet(wb) -> str | None:
    for name in wb.sheetnames:
        if "채권" in name:
            return name
    return wb.sheetnames[0] if wb.sheetnames else None


def _first_present(row: dict, columns: list[str]) -> object | None:
    for col in columns:
        val = row.get(col)
        if val not in (None, ""):
            return val
    return None


def _to_float(value: object, default: float = 0.0) -> float:
    try:
        return float(value) if value not in (None, "") else default
    except (TypeError, ValueError):
        return default


def _parse_bond_row(row: dict, index: int) -> dict | None:
    try:
        remaining_days = int(_to_float(row.get("잔존일수")))
    except (TypeError, ValueError):
        return None
    if remaining_days <= 0:
        return None
    years = remaining_days / 365.0

    notional = _to_float(row.get("결제장부수량(만)")) * 10000
    if notional <= 0:
        return None

    evaluation_amount = _to_float(row.get("평가금액"))
    duration = _to_float(row.get("듀레이션"))

    raw_pvbp = row.get("Duration가중 평가액")
    if raw_pvbp not in (None, ""):
        pvbp = _to_float(raw_pvbp) / 10000.0
    else:
        pvbp = duration * evaluation_amount / 10000.0

    name = str(row.get("종목명") or "").strip()

    # 채권 정적 파라미터 (ParsedPositionOut의 issue_date/coupon_rate/rating
    # 블록). 날짜는 둘 다 Optional 그대로: 정말 비어 있는 행은 None으로 남고
    # (blank-not-zero), 검증 에러가 되지 않는다. issue_date는 start_date에도
    # 복사한다 -- 프런트가 둘 중 무엇을 읽어도 채권 발행일이 나오도록.
    issue_date = _to_text_date(row.get("발행일자"))
    maturity_date = _to_text_date(row.get("만기일자"))
    raw_coupon = row.get("표면이율")
    try:
        coupon_rate = float(raw_coupon) if raw_coupon not in (None, "") else None
    except (TypeError, ValueError):
        coupon_rate = None
    rating = str(row.get("신용등급") or "").strip() or None

    return {
        "instrument_type": "bond",
        "position_id": name or f"bond-{index}",
        "sector": _map_bond_sector(row.get("상품소분류명")),
        "book": _map_book(row.get("펀드명")),
        "notional": notional,
        "evaluation_amount": evaluation_amount,
        "remaining_days": remaining_days,
        "tenor_bucket": _bucket_for_years(years),
        "entry_yield": _to_float(row.get("매수수익율")),
        "mtm_yield": _to_float(row.get("민평수익율")),
        "duration": duration,
        "pvbp": pvbp,
        "start_date": issue_date,
        "maturity_date": maturity_date,
        "issue_date": issue_date,
        "coupon_rate": coupon_rate,
        # payment_frequency stays unset: inferring 2 vs 4 from the sector is
        # a booking convention the owner has to sign off on, not a fact the
        # sheet carries.
        "rating": rating,
        # Bonds aren't booked as swaps -- kept None (not omitted) so every
        # row in the mixed list has the same key set.
        "fixed_rate": None,
        "pay_fixed": None,
        "float_spread": None,
    }


def _parse_irs_row(row: dict, index: int) -> dict | None:
    name = str(row.get("종목명") or "").strip()
    maturity_date = _to_date(row.get("만기일"))
    if maturity_date is None:
        return None

    raw_notional = _to_float(row.get("현재액면(원화)"))
    notional = abs(raw_notional)
    if notional <= 0:
        return None

    # "수취" (receive fixed) -> pay_fixed=False; anything else ("지급"/pay)
    # -> pay_fixed=True (:302) -- then flip for mirror/rolled trades booked
    # with a negative face value in the source ledger (:304-307).
    pay_fixed = str(row.get("지급 수취") or "").strip() != "수취"
    if raw_notional < 0:
        pay_fixed = not pay_fixed

    fixed_rate_pct = _first_present(row, _FIXED_RATE_COLUMNS)
    try:
        fixed_rate = float(fixed_rate_pct) / 100.0 if fixed_rate_pct is not None else None
    except (TypeError, ValueError):
        fixed_rate = None
    if fixed_rate is None:
        return None

    # Start-date priority, mirroring the source exactly (:317-341): trade
    # date + 1 day (never rolls, most reliable anchor) -> a date embedded in
    # the position name -> the sheet's own start-date column (may have
    # rolled forward for already-reset periods, so it's the last resort).
    start_date = None
    trade_date = _to_date(row.get("최초거래일"))
    if trade_date is not None:
        start_date = trade_date + timedelta(days=1)
    if start_date is None:
        start_date = _parse_name_start_date(name)
    if start_date is None:
        start_date = _to_date(row.get("시작일"))
    if start_date is None:
        return None

    # 기초자산1 containing "CD" -> IRS (CD91D-referenced), else OIS
    # (ExcelUploader.tsx:299).
    sector = "IRS" if "CD" in str(row.get("기초자산1") or "") else "OIS"

    return {
        "instrument_type": "irs",
        "position_id": name or f"portfolio-{index}",
        "sector": sector,
        "book": _map_book(row.get("PORTFOLIO명")),
        "start_date": start_date,
        "maturity_date": maturity_date,
        "notional": notional,
        "fixed_rate": fixed_rate,
        "pay_fixed": pay_fixed,
        "float_spread": 0.0,
        # Bond-only fields, kept None for a uniform key set across rows --
        # PVBP for IRS rows is computed separately (real delta engine, not
        # available at parse time).
        "evaluation_amount": None,
        "remaining_days": None,
        "tenor_bucket": None,
        "entry_yield": None,
        "mtm_yield": None,
        "duration": None,
        "pvbp": None,
        "issue_date": None,
        "coupon_rate": None,
        "rating": None,
    }


def _parse_sheet(wb, sheet_name: str, row_parser) -> list[dict]:
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    parsed_rows: list[dict] = []
    for i, raw_row in enumerate(rows_iter):
        if all(v is None for v in raw_row):
            continue
        parsed = row_parser(dict(zip(header, raw_row)), i)
        if parsed is not None:
            parsed_rows.append(parsed)
    return parsed_rows


def _parse_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        positions: list[dict] = []

        irs_sheet = _find_irs_sheet(wb)
        if irs_sheet is not None:
            positions.extend(_parse_sheet(wb, irs_sheet, _parse_irs_row))

        bond_sheet = _find_bond_sheet(wb)
        if bond_sheet is not None and bond_sheet != irs_sheet:
            positions.extend(_parse_sheet(wb, bond_sheet, _parse_bond_row))
    finally:
        wb.close()
    if not positions:
        raise ValueError(f"{XLSX_NAME}에서 유효한 포지션 데이터를 찾을 수 없습니다.")
    n_bonds = sum(1 for p in positions if p["instrument_type"] == "bond")
    n_irs = len(positions) - n_bonds
    logger.info("parsed %d positions (%d bond, %d irs) from %s", len(positions), n_bonds, n_irs, path.name)
    return positions


def parse(data_dir: Path | str) -> list[dict]:
    """Parse Portfolio Data.xlsx into a list of mixed bond/IRS position
    dicts (see module docstring for shape). Also serves as this loader's
    upload-validation entry point: raises ValueError if the file can't be
    parsed (no usable sheets, no valid rows, etc). Not cached -- unlike the
    market-data loaders, this is parsed once at upload time, not read
    repeatedly per pricing call."""
    path = Path(data_dir) / XLSX_NAME
    return _parse_rows(path)


def summary(data_dir: Path | str) -> dict:
    """Position-count / maturity-range summary for the upload endpoint's status response."""
    positions = parse(data_dir)
    irs_maturities = sorted(p["maturity_date"] for p in positions if p.get("maturity_date"))
    return {
        "positions": len(positions),
        "bonds": sum(1 for p in positions if p["instrument_type"] == "bond"),
        "irs": sum(1 for p in positions if p["instrument_type"] == "irs"),
        "min_maturity": irs_maturities[0] if irs_maturities else None,
        "max_maturity": irs_maturities[-1] if irs_maturities else None,
    }
