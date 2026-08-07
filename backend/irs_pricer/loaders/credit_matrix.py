"""
Load credit-spread curve data from "Credit Matrix Data.xlsx" (Infomax export).

Layout: rows 0-2 are headers (title/metadata row, then a per-block category
title row, then a per-block tenor sub-header row), data starts at row 3,
most recent date first -- same header-row convention as True Data.xlsx and
BOK Base Rate.xlsx. Unlike those two, this sheet is a wide matrix: a
17-column template (1 category-label column + 16 tenor columns) repeated
across the whole row width, one repetition per credit category/rating
(confirmed directly against the workbook: 1292 columns = 76 x 17). Each
block's category label lives in row index 1 at the block's own start
column; row index 2 repeats the same 16 tenor labels for every block, read
once from the first block rather than hardcoded so a source-format tweak
(added/renamed/reordered tenor) can't silently desync this loader from the
sheet.

Rates are percentage (e.g. 2.76 = 2.76%); divided by 100 for consistency
with true_data.py/base_rate.py's decimal-rate convention.
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path

import openpyxl

from .cache import get_cached

logger = logging.getLogger(__name__)

XLSX_NAME = "Credit Matrix Data.xlsx"

_HEADER_ROWS = 3
_LABEL_ROW = 1        # 0-based row index carrying each block's category name
_SUBHEADER_ROW = 2    # 0-based row index carrying the repeating tenor labels
_COL_DATE = 0
_BLOCK_WIDTH = 17      # 1 category-label column + 16 tenor columns

_BlockLayout = tuple[str, int, list[str]]  # (category_label, block_start_col, tenor_labels)


def _row_date(row: tuple) -> date | None:
    val = row[_COL_DATE]
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _read_block_layout(ws) -> list[_BlockLayout]:
    """Return [(category_label, block_start_col, [tenor_label, ...]), ...]
    by reading row 1 (category titles) and row 2 (tenor sub-header,
    identical in every block -- read once from the first block) directly
    from the sheet, rather than hardcoding the 76 category names."""
    header_rows = list(ws.iter_rows(min_row=_LABEL_ROW + 1, max_row=_SUBHEADER_ROW + 1, values_only=True))
    label_row, subheader_row = header_rows[0], header_rows[1]
    total_cols = len(subheader_row)

    tenor_labels = [
        str(v).strip() if v is not None else f"col{offset}"
        for offset, v in enumerate(subheader_row[1:_BLOCK_WIDTH], start=1)
    ]

    blocks: list[_BlockLayout] = []
    for start_col in range(0, total_cols, _BLOCK_WIDTH):
        label = label_row[start_col]
        if label is None:
            continue
        blocks.append((str(label).strip(), start_col, tenor_labels))
    if not blocks:
        raise ValueError(f"{XLSX_NAME}: {_LABEL_ROW + 1}행에서 신용 카테고리 블록을 찾을 수 없습니다.")
    return blocks


def _parse_rows(path: Path) -> tuple[list[_BlockLayout], list[tuple]]:
    """Parse the block layout once, then the dated data block, stopping at
    the first blank-date row once real data has been seen -- same
    contiguous-block-then-padding early break as true_data.py/base_rate.py
    (this sheet is 42MB+; the padding rows dominate a naive full scan)."""
    logger.info("opening %s for parsing…", path.name)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        blocks = _read_block_layout(ws)
        rows: list[tuple] = []
        seen_data = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < _HEADER_ROWS:
                continue
            if _row_date(row) is None:
                if seen_data:
                    logger.debug("early break at physical row %d — %d data rows collected", i + 1, len(rows))
                    break  # contiguous data block ended -- skip the padding rows
                continue
            seen_data = True
            rows.append(row)
    finally:
        wb.close()
    if not rows:
        logger.error("_parse_rows returned 0 rows from %s — check header-row layout", path.name)
    else:
        logger.info("parsed %d rows across %d categories from %s", len(rows), len(blocks), path.name)
    return blocks, rows


def _index_rows(parsed: tuple[list[_BlockLayout], list[tuple]]) -> dict[date, dict[str, dict[str, float]]]:
    """Build a {valuation_date: {category: {tenor: rate}}} lookup once per
    parsed workbook."""
    blocks, rows = parsed
    indexed: dict[date, dict[str, dict[str, float]]] = {}
    for row in rows:
        row_date = _row_date(row)
        if row_date is None:
            continue
        by_category: dict[str, dict[str, float]] = {}
        for label, start_col, tenor_labels in blocks:
            tenors: dict[str, float] = {}
            for offset, tenor_label in enumerate(tenor_labels, start=1):
                val = row[start_col + offset]
                # Not-yet-published intraday points on the most recent row
                # are sometimes a text placeholder ("-" etc) rather than
                # blank -- skip anything that isn't actually numeric instead
                # of letting it crash the whole file.
                if isinstance(val, (int, float)):
                    tenors[tenor_label] = val / 100.0
            if tenors:
                by_category[label] = tenors
        indexed[row_date] = by_category
    return indexed


def _load_indexed_rows(data_dir: Path | str) -> dict[date, dict[str, dict[str, float]]]:
    """Cached, date-indexed view of the parsed workbook -- O(1) lookup by
    valuation_date instead of an O(N) linear scan per lookup."""
    path = Path(data_dir) / XLSX_NAME
    return get_cached(path, lambda p: _index_rows(_parse_rows(p)), cache_key_suffix="by_date")


def common_dates_xlsx(data_dir: Path | str) -> list[date]:
    """Return all valuation dates present in the workbook, sorted ascending.

    Also serves as this loader's upload-validation entry point: raises
    ValueError if the file can't be parsed as a Credit Matrix workbook
    (wrong header layout, no dated rows, etc.)."""
    dates = _load_indexed_rows(data_dir).keys()
    if not dates:
        raise ValueError(f"{XLSX_NAME}에서 날짜 데이터를 찾을 수 없습니다.")
    return sorted(dates)


def category_series(
    data_dir: Path | str,
    category: str,
    tenor: str,
    start_date: date | None = None,
    end_date: date | None = None,
) -> list[tuple[date, float]]:
    """One category+tenor's decimal-rate time series, ascending by date and
    optionally clipped to [start_date, end_date]. Reads the already-cached
    full {date: {category: {tenor: rate}}} structure (O(1) after first parse,
    no 42MB re-read -- see cache.py) and filters in-memory. Dates where that
    category/tenor has no value are simply absent from the result."""
    indexed = _load_indexed_rows(data_dir)
    out: list[tuple[date, float]] = []
    for d in sorted(indexed.keys()):
        if start_date is not None and d < start_date:
            continue
        if end_date is not None and d > end_date:
            continue
        val = indexed[d].get(category, {}).get(tenor)
        if val is not None:
            out.append((d, val))
    return out


def curve_at(
    data_dir: Path | str,
    category: str,
    val_date: date | None = None,
) -> dict[str, float]:
    """One category's whole curve as {raw_tenor: decimal_rate}, each tenor
    carrying its latest value at or before val_date (latest overall when
    val_date is None).

    Equivalent to calling category_series() once per tenor and taking each
    one's last point, but walks the date index a single time instead of
    re-sorting it per tenor -- the curve for a (category, val_date) pair is
    identical for every bond sharing that sector+rating, so callers pricing a
    book would otherwise rebuild the same curve once per position.

    A tenor absent on the newest date keeps its own last known value, which is
    exactly what the per-tenor series[-1] lookup yields: gaps are skipped, not
    zero-filled.
    """
    indexed = _load_indexed_rows(data_dir)
    out: dict[str, float] = {}
    for d in sorted(d for d in indexed if val_date is None or d <= val_date):
        # Ascending, so a later date overwrites an earlier one and each tenor
        # ends on its own latest value.
        out.update(indexed[d].get(category, {}))
    return out


def categories(data_dir: Path | str) -> set[str]:
    """Every category label present across the workbook -- lets a caller
    validate a taxonomy's raw-category strings against the live file."""
    indexed = _load_indexed_rows(data_dir)
    return {cat for by_cat in indexed.values() for cat in by_cat}


def summary(data_dir: Path | str) -> dict:
    """Row/category/date-range summary for the upload endpoint's status response."""
    indexed = _load_indexed_rows(data_dir)
    dates = sorted(indexed.keys())
    categories = {cat for by_cat in indexed.values() for cat in by_cat}
    return {
        "rows": len(dates),
        "categories": len(categories),
        "min_date": dates[0] if dates else None,
        "max_date": dates[-1] if dates else None,
    }


def daily_shift(data_dir: Path | str) -> dict[str, dict[str, float]]:
    """Return the day-over-day shift in basis points (latest row - previous row)
    for each credit category and tenor."""
    indexed = _load_indexed_rows(data_dir)
    if len(indexed) < 2:
        return {}

    dates = sorted(indexed.keys(), reverse=True)
    latest_row = indexed[dates[0]]
    prev_row = indexed[dates[1]]

    shift: dict[str, dict[str, float]] = {}

    for category, latest_tenors in latest_row.items():
        prev_tenors = prev_row.get(category, {})
        cat_shift: dict[str, float] = {}
        for tenor, val_latest in latest_tenors.items():
            val_prev = prev_tenors.get(tenor)
            if val_prev is not None:
                # _load_indexed_rows stores true decimals (e.g. 0.0276 for 2.76%)
                # Multiply decimal difference by 10000 to get basis points
                cat_shift[tenor] = (val_latest - val_prev) * 10000.0
        if cat_shift:
            shift[category] = cat_shift

    return shift
