"""The BOK base rate — the policy anchor every % chart is drawn against.

[OWNER, 2026-07-31] CD and the base rate are always drawn together, and the
base rate rides along on every %-unit chart (outrights + forwards). Spread,
butterfly and volatility charts are bp / ratio and are excluded: a 2.75 line
on a ±30bp axis is not a comparison, it is a rescale.

Two things make this file more than a second `dataset.py`:

**It is a STEP, not a series.** The rate changes only on a Monetary Policy
Board decision and holds flat in between, so it is drawn with square corners
and read as a level that was *in force* on a date, never interpolated.

**Carrying it forward is a claim.** Holding the last observed rate out to the
IRS as-of date is correct exactly when no Board meeting fell in the gap — and
silently wrong when one did, in the one direction that matters: it would draw
the OLD rate through the day it changed, on every chart at once, with nothing
on screen to say so.

**The source is ECOS, the publisher itself** [OWNER, 2026-09-01]. It used to be
`data/bokbaserate.xlsx`, an Infomax export refreshed by hand — which meant the
guard below fired every time a meeting landed between two exports, and the step
stopped short until somebody re-exported. `app/funding.py` had already moved to
ECOS for the same reason on 2026-08-20 (see `app/ecos.py`); this module was the
half that stayed behind, and the 2026-08-27 hike is what surfaced it.

The workbook is kept as a FALLBACK — no key, no network and no cache means the
charts still stand on the last hand snapshot, and `warnings` says so. The guard
below is unchanged: ECOS can be stale too (it publishes on a lag), and a fresh
source is not the same fact as a source that has seen the meeting.

So the carry is bounded by the meeting calendar. If a meeting sits between the
source's last date and the dataset's as-of date, the step **ends at the last
date it can vouch for** and a warning is recorded. The line visibly stops
short rather than asserting a rate nobody checked; a missing tail is a
question the reader can see, a fabricated one is not.
"""

from __future__ import annotations

import datetime as dt
import json
import logging
from bisect import bisect_right
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

log = logging.getLogger(__name__)

# Sheet geometry — the Infomax add-in's export, same shape as irsdata.xlsx:
# row 1 metadata, row 2 the code (`한국:기준금리`) + unit, row 3 field names,
# then data. Unlike the IRS workbook this one is DESCENDING (newest first).
FIRST_DATA_ROW = 4
DATE_COL = 0
VALUE_COL = 1

# Plausible band for a policy rate in percent. A nonsense check, not a view.
RATE_MIN_PCT = -1.0
RATE_MAX_PCT = 15.0

# Monetary Policy Board rate-decision dates. A SECOND COPY of the `mpc`
# entries in `frontend/src/data/calendar.json`, which is the owner-verified
# original. It is copied rather than imported because the backend must not
# reach into the frontend source tree at runtime, and
# `tests/test_policy.py::test_mpc_dates_match_the_calendar` fails if the two
# ever disagree — so the duplication cannot rot silently.
MPC_DATES = [
    dt.date(2026, 1, 15),
    dt.date(2026, 2, 26),
    dt.date(2026, 4, 10),
    dt.date(2026, 5, 28),
    dt.date(2026, 7, 16),
    dt.date(2026, 8, 27),
    dt.date(2026, 10, 22),
    dt.date(2026, 11, 26),
]

# V2-LOCAL: v1 은 `frontend/src/data/`, v2 는 프론트가 리포 루트라 `src/data/`.
# 이 파일은 v2 에 아직 없고(그래서 아래 함수는 None 을 돌려준다) 그건 정상이다 —
# 테스트 보조이지 런타임이 아니다. 다만 경로는 v2 의 실제 자리를 가리켜야, 파일이
# 생긴 날 조용히 안 읽히는 일이 없다.
CALENDAR_JSON = (
    Path(__file__).resolve().parents[2] / "src" / "data" / "calendar.json"
)


def mpc_dates_from_calendar() -> list[dt.date] | None:
    """The `mpc` dates in the frontend's verified calendar, or None if the file
    is not there (a backend-only checkout). Test support, not runtime."""
    if not CALENDAR_JSON.exists():
        return None
    doc = json.loads(CALENDAR_JSON.read_text(encoding="utf-8"))
    return sorted(
        dt.date.fromisoformat(e["date"])
        for e in doc["events"]
        if e.get("kind") == "mpc"
    )


class PolicyFileError(Exception):
    """The workbook cannot be trusted at all — a wrong number, not an old one."""


@dataclass
class BaseRate:
    dates: list[dt.date]          # ascending, one per observation in the file
    values: list[float]           # percent, aligned to `dates`
    warnings: list[str] = field(default_factory=list)
    #: Where these numbers came from, in words the warning can print. There are
    #: two sources now (ECOS, and the hand workbook as fallback) and the carry
    #: guard has to name the one it is actually bounded by — otherwise it tells
    #: the reader to refresh a file that is not being read.
    source: str = "?"

    @property
    def asof(self) -> dt.date:
        return self.dates[-1]

    @property
    def latest(self) -> float:
        return self.values[-1]

    def at(self, date: dt.date) -> float | None:
        """The rate IN FORCE on `date` — the last decision at or before it.
        None before the file's first observation; never interpolated."""
        i = bisect_right(self.dates, date)
        return self.values[i - 1] if i > 0 else None


def load_base_rate(xlsx_path: Path) -> BaseRate:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    next(rows)                 # row 1: metadata
    next(rows)                 # row 2: code + unit
    fields = next(rows)        # row 3: field names
    if fields[DATE_COL] != "일자":
        raise PolicyFileError(
            f"{xlsx_path.name}: expected 일자 in the first field column, "
            f"found {fields[DATE_COL]!r} — is this the base-rate export?"
        )

    pairs: list[tuple[dt.date, float]] = []
    for n, row in enumerate(rows, start=FIRST_DATA_ROW):
        raw_date, raw_value = row[DATE_COL], row[VALUE_COL]
        if raw_date is None:
            continue           # the export pads to a fixed row count
        if not isinstance(raw_date, dt.datetime):
            raise PolicyFileError(f"{xlsx_path.name} row {n}: bad date {raw_date!r}")
        if raw_value is None:
            continue
        value = float(raw_value)
        if not RATE_MIN_PCT <= value <= RATE_MAX_PCT:
            raise PolicyFileError(
                f"{xlsx_path.name} row {n}: {value} is not a policy rate in percent"
            )
        pairs.append((raw_date.date(), value))
    wb.close()

    if not pairs:
        raise PolicyFileError(f"{xlsx_path.name}: no observations")

    pairs.sort(key=lambda p: p[0])   # the export is newest-first
    return BaseRate(
        dates=[d for d, _v in pairs],
        values=[v for _d, v in pairs],
        source=xlsx_path.name,
    )


def load_base_rate_ecos() -> BaseRate:
    """The same step, from 722Y001/D/0101000 (1999-05-06 ~).

    ⚠ UNITS. `ecos.base_rate_series()` returns DECIMALS (0.03) because that is
    what `funding` wants; `BaseRate.values` is PERCENT (3.00) because that is
    what the charts draw. The x100 below is the whole difference, and getting
    it wrong produces a plausible-looking number rather than an exception —
    this repo has already paid for that once, on the carry term. The band check
    is what catches it: 0.03 is not a policy rate in percent.
    """
    from . import ecos

    # `round` 는 표시용 반올림이 아니라 **왕복 잡티 제거**다. ECOS 는 연% 를
    # 소수로 나눠 두었으므로 x100 이 1.75 를 1.7500000000000002 로 만들고,
    # `decisions()` 가 코너를 정확한 같음(`!=`)으로 가리는 이상 그 잡티는
    # 화면까지 그대로 간다. 정책금리는 0.25 배수라 6자리 반올림이 정확하다.
    pairs = [(d, round(v * 100.0, 6)) for d, v in ecos.base_rate_series()]
    if not pairs:
        raise PolicyFileError("ECOS: no observations")
    for d, v in pairs:
        if not RATE_MIN_PCT <= v <= RATE_MAX_PCT:
            raise PolicyFileError(f"ECOS {d}: {v} is not a policy rate in percent")
    return BaseRate(
        dates=[d for d, _v in pairs],
        values=[v for _d, v in pairs],
        source=f"ECOS {ecos.BASE_RATE_STAT}",
    )


def load_base_rate_auto(xlsx_path: Path) -> BaseRate:
    """ECOS first, the hand workbook if ECOS cannot answer at all.

    The fallback is not silent — it lands in `warnings`, which `policy_step`
    passes through to the payload, because a chart drawn on a hand snapshot is
    a different fact from one drawn on the publisher's own series.
    """
    try:
        return load_base_rate_ecos()
    except Exception as exc:                     # EcosError, key missing, parse
        log.warning("[policy] ECOS 를 못 읽어 워크북으로 물러섭니다 — %s", exc)
        base = load_base_rate(xlsx_path)
        base.warnings.append(
            f"[policy] base rate came from {xlsx_path.name} (hand export, "
            f"last {base.asof}) because ECOS could not answer: {exc}"
        )
        return base


def decisions(base: BaseRate) -> list[tuple[dt.date, float]]:
    """The step's CORNERS: the first date, plus every date the rate changed.
    Two thousand daily rows describe about a dozen decisions; the chart only
    needs the corners, and sending the rest would be sending the same number
    over and over (§20)."""
    out = [(base.dates[0], base.values[0])]
    for d, v in zip(base.dates[1:], base.values[1:]):
        if v != out[-1][1]:
            out.append((d, v))
    return out


def policy_step(base: BaseRate, asof: dt.date) -> dict:
    """The payload the charts draw: step corners, plus the date the step is
    entitled to run to.

    `through` is `asof` when the carry is safe and the workbook's own last
    date when it is not — see the module docstring. Consumers draw the final
    level flat out to `through` and STOP; they must not extend to their own
    axis end, which is the mistake this field exists to prevent.
    """
    warnings: list[str] = list(base.warnings)   # 출처가 워크북이면 그 사실부터
    through = asof
    if base.asof < asof:
        missed = [d for d in MPC_DATES if base.asof < d <= asof]
        if missed:
            through = base.asof
            warnings.append(
                f"[policy] base rate is stale to {base.asof} and the Board met "
                f"{', '.join(d.isoformat() for d in missed)} — the step ends at "
                f"{base.asof} rather than carrying an unverified rate to {asof}. "
                f"Refresh the source ({base.source})."
            )
    for w in warnings:
        log.warning(w)
    return {
        "unit": "%",
        "asof": base.asof.isoformat(),
        "through": through.isoformat(),
        "steps": [
            {"date": d.isoformat(), "rate": v}
            for d, v in decisions(base)
            if d <= through
        ],
        "latest": base.at(through),
        # V2-LOCAL, 2026-08-14 — 앞으로 남은 금통위 날짜. 시뮬레이션의 기준금리
        # 이벤트가 이걸 읽어 날짜 칸을 채운다.
        #
        # 프론트에 두 번째 목록을 두지 않는 이유는 이 모듈이 이미 겪은 것이다:
        # `MPC_DATES` 자체가 `calendar.json` 의 사본이고, 그 중복이 조용히 썩지
        # 않도록 테스트가 둘을 대조한다. 세 번째 사본을 브라우저에 두면 그
        # 대조에서 빠진 사본이 하나 생긴다.
        #
        # `asof` 이후만 보낸다 — 지난 회의는 이미 `steps` 에 결과로 들어 있고,
        # 시뮬레이션이 놓을 수 있는 것은 아직 안 온 회의뿐이다.
        "upcoming": [d.isoformat() for d in MPC_DATES if d > asof],
        "warnings": warnings,
    }
