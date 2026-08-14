"""Load data/irsdata.xlsx into an in-memory store of daily par-rate series.

Sheet layout (Infomax export):
  row 1: query metadata (start/end/count/...)
  row 2: series codes, merged cells ("원화 IRS 종합코드 6개월", ..., "...콜금리")
  row 3: field names ("일자", "MID종가" x13, "수익률")
  row 4+: data rows, dates DESCENDING

Columns B..N are the 13 IRS par tenors, column O is the call rate (1D).
Values are percent (e.g. 4.135 = 4.135%).
"""

from __future__ import annotations

import datetime as dt
import logging
import re
from bisect import bisect_left
from dataclasses import dataclass, field
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)


class DataFileError(ValueError):
    """The workbook cannot be trusted, so the server does not start.

    Raised for the failures Pass A classified PLAUSIBLE-WRONG: a duplicated or
    out-of-order date silently misdirects every basis lookup, and an unbounded
    value (a decimal slip) flows into the bootstrap and out through every
    derived number. Both used to load without a word.

    The message always names the CELL — `D57`, not "row 57" and not "the 3Y
    series" — because the person fixing this is looking at a spreadsheet, and
    a cell reference is what they can type into the name box.
    """


# Sheet geometry: metadata, merged labels, field names, then data.
FIRST_DATA_ROW = 4

# Plausible band for a KRW par rate in percent. Wide on purpose — this is a
# nonsense check, not a view. It catches a decimal slip (4135 for 4.135) and a
# sign error, and would not fire on any rate this market has printed.
RATE_MIN_PCT = -5.0
RATE_MAX_PCT = 25.0

# The market's timezone. "Which day is today" is a question about Seoul —
# the data is KRW closes — not about wherever this process happens to run.
# Via the tz database, not a fixed +9, for the same reason freshness.ts
# gives: facts about the world belong in the platform's tzdata.
MARKET_TZ = ZoneInfo("Asia/Seoul")


def market_today() -> dt.date:
    """Today's calendar date in Seoul, at this instant."""
    return dt.datetime.now(MARKET_TZ).date()


# Calendar days between consecutive observations before the file is called
# stale. Four-day weekends and Chuseok/Seollal clusters are ordinary; ten days
# means the update stopped. STALE, not unusable — the numbers are still true,
# they are just old, and refusing to start would be the wrong answer.
MAX_GAP_DAYS = 10


def _cell(row_no: int, col: int) -> str:
    """`(57, 3)` → `'D57'`. Columns are 0-based here, 1-based in the sheet."""
    return f"{get_column_letter(col + 1)}{row_no}"

# Wall node order. 3M (CD91) is in the product spec but ABSENT from the
# current data export — the loader records it in `missing_nodes` instead of
# faking it. Keep this list in spec order so consumers never re-sort.
SPEC_NODE_ORDER = [
    "1D", "3M", "6M", "9M", "1Y", "1.5Y", "2Y", "3Y", "5Y", "10Y",
]

# Volatility rows — one relative-ATR ratio per tenor. Its OWN list, deliberately
# not DISPLAY_TENORS (2026-07-31): the two were the same six by coincidence, and
# widening the derived universe for a 6M/9M butterfly would otherwise have grown
# the 변동성 tab as a silent side effect of an unrelated decision. Nobody asked
# for 6M/9M volatility rows; if that is wanted it is a separate ruling here.
VOL_TENORS = ["1Y", "1.5Y", "2Y", "3Y", "5Y", "10Y"]

# Display tenor set for spreads/flies. [OWNER] Widened to include 6M and 9M
# on 2026-07-31, because 6M/9M/1Y is one of the four 주요 버터플라이 and 주요
# has to be a SUBSET of 전체 for the divider between them to mean anything.
# The combinatorics are quadratic/cubic in this list — 6→8 tenors takes
# spreads 15→28 and flies 20→56 — so do not widen it casually.
DISPLAY_TENORS = ["6M", "9M", "1Y", "1.5Y", "2Y", "3Y", "5Y", "10Y"]

# Tenor id → years, for explicit numeric sort keys (§6/§16). Unknown → +inf so
# a genuinely unmapped tenor sorts to the end loudly, never silently mid-list.
TENOR_YEARS: dict[str, float] = {
    "1D": 1.0 / 365.0, "3M": 0.25, "6M": 0.5, "9M": 0.75, "1Y": 1.0,
    "1.5Y": 1.5, "2Y": 2.0, "3Y": 3.0, "4Y": 4.0, "5Y": 5.0, "6Y": 6.0,
    "7Y": 7.0, "8Y": 8.0, "9Y": 9.0, "10Y": 10.0,
}


def tenor_years(tenor: str) -> float:
    return TENOR_YEARS.get(tenor, float("inf"))


# Live-quoted curve nodes (the actual node set); every other tenor is
# interpolated (§6). The quoted/interpolated dot marker reads this.
QUOTED_NODES = frozenset(
    {"1D", "3M", "6M", "9M", "1Y", "1.5Y", "2Y", "3Y", "5Y", "10Y"}
)


def _tenor_id(label: str) -> str:
    """Map a Korean series label to a tenor id like '6M', '1.5Y', '1D'."""
    if "콜금리" in label:
        return "1D"
    if "CD" in label:
        return "3M"  # CD 91d average — the spec's 3M node (IRS 3M = CD91)
    m = re.search(r"(\d+)개월", label)
    if m:
        months = int(m.group(1))
        if months % 12 == 0:
            return f"{months // 12}Y"
        if months == 18:
            return "1.5Y"
        return f"{months}M"
    m = re.search(r"(\d+)년", label)
    if m:
        return f"{m.group(1)}Y"
    raise ValueError(f"unrecognized series label: {label!r}")


@dataclass
class Dataset:
    dates: list[dt.date]                       # ascending
    series: dict[str, list[float | None]]      # tenor id -> values aligned to dates
    tenor_order: list[str]                     # as found in the sheet, 1D first
    missing_nodes: list[str] = field(default_factory=list)
    # Things wrong with the file that do NOT make it untrustworthy: gaps,
    # blanks, an old last observation. STALE is not UNUSABLE — these are
    # logged at startup and the server runs. Anything that would make a
    # displayed number wrong raises DataFileError instead.
    warnings: list[str] = field(default_factory=list)
    # 이 데이터셋이 어디서 왔는가 [OWNER, 2026-08-11 — "엑셀데이터에 연결되어
    # 있다고 말은 해줘야 해"]. 값은 넷뿐이다:
    #   "sql"          mkt_irs_close 그대로
    #   "sql+xlsx-1d"  SQL 이되, asof 의 1D 만 엑셀에서 채움
    #   "sql+xlsx-day" SQL 이되, asof 하루 전체를 엑셀에서 덧붙임
    #   "xlsx"         SQL 을 못 읽어 엑셀 전체로 폴백
    # 화면(freshness 칩)과 manifest 가 이 값을 그대로 내보낸다. 라벨이지 판정이
    # 아니다 — 어떤 값이든 서버는 뜬다.
    source: str = "sql"
    # 파생 페이로드 디스크 캐시의 키. 병합 로더만 채운다 — 워터마크(sql)나
    # 파일 바이트(xlsx)에 병합분 지문이 붙는다. 빈 문자열이면 호출자가
    # sql_data_hash 로 직접 만든다 (단일 출처 경로).
    data_key: str = ""

    @property
    def asof(self) -> dt.date:
        return self.dates[-1]

    def latest(self, tenor: str) -> float | None:
        return self.series[tenor][-1]


def load_dataset(xlsx_path: Path, today: dt.date | None = None) -> Dataset:
    """Load the workbook. `today` (Seoul date, defaulting to the clock) is the
    전일종가 cutoff — see the drop below; tests inject it to stay date-free."""
    today = today or market_today()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    next(rows)                 # row 1: metadata
    labels = next(rows)        # row 2: codes (merged -> None gaps)
    fields = next(rows)        # row 3: field names

    if fields[0] != "일자":
        raise DataFileError(
            f"A3: expected the date column header '일자', found {fields[0]!r} "
            f"— is this the Infomax export, with its metadata row on top?"
        )

    # Resolve merged-cell label gaps: a None label belongs to the previous
    # cell's series (the code cell is merged across two columns).
    n_cols = len(fields)
    col_tenors: dict[int, str] = {}
    prev_label: str | None = None
    for col in range(n_cols):
        raw = labels[col] if col < len(labels) else None
        if raw is not None:
            prev_label = str(raw)
        if col == 0:
            continue  # date column
        if prev_label is None:
            raise DataFileError(f"{_cell(2, col)}: no series label")
        try:
            col_tenors[col] = _tenor_id(prev_label)
        except ValueError as exc:
            raise DataFileError(f"{_cell(2, col)}: {exc}") from None

    tenor_order = [col_tenors[c] for c in sorted(col_tenors)]
    if len(set(tenor_order)) != len(tenor_order):
        dupes = sorted({t for t in tenor_order if tenor_order.count(t) > 1})
        cols = [
            _cell(2, c) for c in sorted(col_tenors) if col_tenors[c] in dupes
        ]
        raise DataFileError(
            f"two columns carry the same tenor {dupes}: {', '.join(cols)}"
        )

    dates_desc: list[dt.date] = []
    sheet_rows: list[int] = []   # sheet row per observation, for error messages
    values_desc: dict[str, list[float | None]] = {t: [] for t in tenor_order}
    blanks: dict[str, int] = {t: 0 for t in tenor_order}

    for row_no, row in enumerate(rows, start=FIRST_DATA_ROW):
        raw_date = row[0]
        if raw_date is None:
            continue
        if not isinstance(raw_date, dt.datetime):
            raise DataFileError(
                f"{_cell(row_no, 0)}: expected a date, found {raw_date!r}"
            )
        dates_desc.append(raw_date.date())
        sheet_rows.append(row_no)
        for col, tenor in col_tenors.items():
            v = row[col] if col < len(row) else None
            if v is None:
                blanks[tenor] += 1
                values_desc[tenor].append(None)
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                raise DataFileError(
                    f"{_cell(row_no, col)} ({tenor}): expected a number, "
                    f"found {v!r}"
                ) from None
            # A decimal slip is invisible without this: 4135 bootstraps just
            # as happily as 4.135, and every derived number inherits it.
            if not RATE_MIN_PCT <= fv <= RATE_MAX_PCT:
                raise DataFileError(
                    f"{_cell(row_no, col)} ({tenor}): {fv} is outside "
                    f"{RATE_MIN_PCT}%..{RATE_MAX_PCT}% — check for a "
                    f"misplaced decimal point or a sign"
                )
            values_desc[tenor].append(fv)
    wb.close()

    if not dates_desc:
        raise DataFileError("no data rows found")

    # Every date lookup in the product is a bisect on an assumed-ascending,
    # assumed-unique list (`derive.value_at`). A duplicate or a swapped pair
    # does not crash it — it silently returns the wrong row, so D-1/MTD/YTD
    # all read the wrong day while the levels look perfect. This is
    # the single most dangerous thing a hand-updated sheet can do.
    first_seen: dict[dt.date, int] = {}
    for date, row_no in zip(dates_desc, sheet_rows):
        if date in first_seen:
            raise DataFileError(
                f"{_cell(row_no, 0)}: {date} already appears at row "
                f"{first_seen[date]} — every date must be unique"
            )
        first_seen[date] = row_no

    ascending = dates_desc[0] < dates_desc[-1]
    dates = dates_desc if ascending else list(reversed(dates_desc))
    order_rows = sheet_rows if ascending else list(reversed(sheet_rows))
    series = {
        t: (vals if ascending else list(reversed(vals)))
        for t, vals in values_desc.items()
    }

    # Orientation is decided from the first and last row alone, so a swapped
    # pair in the middle survives that check untouched. Check every step.
    for i in range(1, len(dates)):
        if dates[i] <= dates[i - 1]:
            raise DataFileError(
                f"rows {order_rows[i - 1]} and {order_rows[i]}: dates run "
                f"{dates[i - 1]} then {dates[i]} — rows must be in date "
                f"order, with no repeats"
            )

    return _finalize(dates, series, tenor_order, order_rows, today,
                     source="xlsx")


def _finalize(
    dates: list[dt.date],
    series: dict[str, list[float | None]],
    tenor_order: list[str],
    order_rows: list[int],
    today: dt.date,
    source: str = "sql",
) -> Dataset:
    """파싱이 끝난 뒤의 **공통 규칙** — 전일종가 컷, 빈칸·공백 검사, 경고.

    2026-08-07 에 `load_dataset` 의 꼬리에서 떼어냈다. 출처가 둘이 되었기
    때문이다(엑셀 · MySQL). 이 부분이 두 벌이 되면 두 출처가 서로 다른 날을
    "오늘" 로 자르거나 한쪽만 빈 칼럼을 통과시키는 일이 생기고, 그건 화면에서
    구별되지 않는다. 파싱은 출처마다 다르고 **판정은 하나**다.

    `order_rows` 는 오류 문장에 찍히는 행 번호다 — 엑셀은 시트 행, DB 는 정렬된
    결과의 순번이다.
    """
    warnings: list[str] = []

    # ── 전일종가 rule [OWNER, 2026-08-05] ───────────────────────────────────
    # A row dated today is NOT a close: the Infomax export refreshed at 11:00
    # writes 11:00's live quotes into today's row, and they load exactly like
    # a settled close — every "현재" on screen then quietly means "whenever
    # the workbook was last saved". Prices for the current Seoul date are
    # therefore treated as MISSING, whatever the wall clock says: the basis
    # is always the last completed close (asof = the previous business day,
    # d1 = the one before it). A future-dated row falls under the same cut.
    # The cut is HERE, at the single choke point every consumer reads
    # through, so summary/curve/backtest/forwards all shift together.
    cut = bisect_left(dates, today)
    if cut < len(dates):
        dropped = [d.isoformat() for d in dates[cut:]]
        dates = dates[:cut]
        series = {t: vals[:cut] for t, vals in series.items()}
        warnings.append(
            f"dropped {len(dropped)} intraday row(s) on/after {today} "
            f"({', '.join(dropped)}) — 전일종가 rule: today's quotes are "
            "live, not a close"
        )
    if not dates:
        raise DataFileError(
            f"no completed closes: every data row is dated on/after {today}"
        )
    # Blank counts are re-derived from the KEPT rows — the parse-time tallies
    # include any dropped intraday rows, and an all-blank column must be
    # judged on what will actually be served.
    blanks = {
        t: sum(v is None for v in vals) for t, vals in series.items()
    }

    # A column of nothing is not a series. Everything downstream would read
    # `None` forever and show an em dash where a rate belongs.
    for tenor, n in blanks.items():
        if n == len(dates):
            raise DataFileError(f"{tenor}: every value is blank")
        if n:
            warnings.append(f"{tenor}: {n} blank value(s)")

    # STALE, not unusable: say it and carry on.
    gaps = [
        (dates[i] - dates[i - 1]).days
        for i in range(1, len(dates))
    ]
    if gaps and max(gaps) > MAX_GAP_DAYS:
        i = gaps.index(max(gaps)) + 1
        warnings.append(
            f"{max(gaps)}-day gap between {dates[i - 1]} and {dates[i]} "
            f"(row {order_rows[i]}) — the file may have gone unupdated"
        )
    age = (dt.date.today() - dates[-1]).days
    if age > MAX_GAP_DAYS:
        warnings.append(f"last observation {dates[-1]} is {age} days old")

    for w in warnings:
        log.warning("[dataset] %s", w)

    # 1D (call) first, then the sheet's IRS tenor order.
    ordered = sorted(tenor_order, key=lambda t: (t != "1D", tenor_order.index(t)))
    missing = [t for t in SPEC_NODE_ORDER if t not in series]
    return Dataset(dates=dates, series=series, tenor_order=ordered,
                   missing_nodes=missing, warnings=warnings, source=source)


# ── MySQL ─────────────────────────────────────────────────────────────────────
#
# 출처가 DB 로 옮겨간다 [OWNER, 2026-08-07 — "무조건 SQL 쪽이 정답임"].
#
# 대조 결과가 그 판단의 근거다 (backend/scripts/check_mysql.py, 39,220개 값):
#   3M~10Y   불일치 0~1건 / 2,616 — 사실상 완전 일치. 그 "1건" 은 전부 xlsx 의
#            **마지막 행**(2026-08-05)이고, 다른 2,615일이 소수점 끝까지 맞는
#            모양이라 그 행이 종가가 아니라 장중 스냅샷이었다는 뜻이다.
#   1D       불일치 2,105건 / 2,606 (80.8%), 최대 61.4bp — **다른 계열**이다.
#            오너가 SQL 을 정답으로 정했으므로 그대로 받는다. 과거 짧은 끝
#            커브가 달라지고, 그건 의도된 변경이다.
#
# DB 컬럼 → 테너. `_tenor_id()` 가 한글 라벨에 하는 일을 컬럼명에 대해 하는
# 것이고, 매핑이 1:1 이라 함수가 아니라 표다.
SQL_COLUMN_TENOR: dict[str, str] = {
    "call_rate": "1D",   # 콜금리
    "cd_rate": "3M",     # CD 91일 — 스펙의 3M 노드 (IRS 3M = CD91)
    "irs_6m": "6M",
    "irs_9m": "9M",
    "irs_1y": "1Y",
    "irs_18m": "1.5Y",
    "irs_2y": "2Y",
    "irs_3y": "3Y",
    "irs_4y": "4Y",
    "irs_5y": "5Y",
    "irs_6y": "6Y",
    "irs_7y": "7Y",
    "irs_8y": "8Y",
    "irs_9y": "9Y",
    "irs_10y": "10Y",
}


def load_dataset_sql(today: dt.date | None = None) -> Dataset:
    """`mkt_irs_close` 를 읽어 `Dataset` 을 만든다. 엑셀 로더와 **같은 판정**을
    지난다 (`_finalize`) — 전일종가 컷도, 빈 칼럼 거부도, 경고도.

    엑셀 로더와 다른 것은 파싱뿐이다: 시트의 병합 셀·한글 라벨 대신 컬럼 하나가
    테너 하나다. 그래서 여기에는 라벨 해석도, 중복 칼럼 검사도 없다 — 컬럼명이
    유일하다는 것을 DB 가 보장한다.
    """
    from .mysqldb import irs_close_rows  # 지연 import: 엑셀 경로는 DB 를 안 켠다

    today = today or market_today()
    rows = irs_close_rows()            # 날짜 오름차순
    if not rows:
        raise DataFileError("mkt_irs_close 가 비어 있다")

    dates: list[dt.date] = []
    series: dict[str, list[float | None]] = {t: [] for t in SQL_COLUMN_TENOR.values()}
    for r in rows:
        d = r["irs_date"]
        if not isinstance(d, dt.date):
            raise DataFileError(f"irs_date 가 날짜가 아니다: {d!r}")
        dates.append(d)
        for col, tenor in SQL_COLUMN_TENOR.items():
            v = r.get(col)
            series[tenor].append(None if v is None else float(v))

    # 엑셀 로더가 시트 행에 대해 하는 검사와 같은 것 — 정렬은 SQL 이 하지만,
    # 같은 날짜가 두 번 있으면 여기서 잡힌다 (이 테이블에는 PK 가 없다).
    for i in range(1, len(dates)):
        if dates[i] <= dates[i - 1]:
            raise DataFileError(
                f"mkt_irs_close: {dates[i - 1]} 다음에 {dates[i]} — 날짜가 "
                "오름차순이어야 하고 중복이 없어야 한다"
            )

    tenor_order = list(SQL_COLUMN_TENOR.values())
    order_rows = list(range(1, len(dates) + 1))
    return _finalize(dates, series, tenor_order, order_rows, today,
                     source="sql")


# ── 병합: SQL 우선, 엑셀 보충 [OWNER, 2026-08-11] ────────────────────────────
#
# 아침 자동 굽기의 데이터 규칙. 오너 지시 그대로다:
#   "혹시 SQL 데이터가 없다면 엑셀 데이터를 참조하는 방식으로 할 거고" —
#   "만약 1D가 없다면, 1D는 엑셀에서 가져와서 채우는 걸로 하고" —
#   "만약 엑셀데이터를 받아온다면 이건 엑셀데이터에 연결되어있다고 말은
#    해줘야 해"
#
# 그래서 판정은 **기대 전영업일 하루**에 대해서만 내린다. SQL 이 그 날을 온전히
# 들고 있으면 SQL 그대로, 1D 만 비면 그 칸만 엑셀, 그 날이 통째로 없으면 그
# 하루를 엑셀에서 덧붙인다. 과거사(history)는 절대 엑셀로 갈아타지 않는다 —
# 1D 는 두 출처가 **다른 계열**(80.8% 불일치)이라, 폴백이 역사를 바꾸면 SQL 이
# 뒤늦게 적재된 날 1D 차트에 유령 점프가 생긴다. 전체 폴백("xlsx")은 SQL 을
# 아예 못 읽는 비상시 뿐이고, 그때는 칩이 말한다.

#: 병합 로더의 기본 엑셀 경로. 서버가 읽는 정본은 MySQL 이고(2026-08-07),
#: 이 파일은 아침 자동화가 갱신해 두는 **보충 출처**다.
DEFAULT_XLSX = Path(__file__).resolve().parents[2] / "data" / "irsdata.xlsx"


def prev_kr_business_day(today: dt.date) -> dt.date:
    """`today`(서울 날짜) 직전의 한국 영업일 — "기대하는 전일 종가"의 날짜.

    달력은 엔진 것 하나뿐이다 (`engine_port._is_kr_business_day`). 지연
    import 인 이유는 mysqldb 와 같다 — 엑셀만 만지는 경로가 QuantLib 을
    끌어들일 이유가 없다.
    """
    from .engine_port import _is_kr_business_day  # 지연 import

    d = today - dt.timedelta(days=1)
    while not _is_kr_business_day(d):
        d -= dt.timedelta(days=1)
    return d


def _xlsx_value_at(xlsx_ds: Dataset, tenor: str, date: dt.date) -> float | None:
    """엑셀 데이터셋에서 특정 날짜의 값 하나. 없으면 None — 예외가 아니다."""
    if tenor not in xlsx_ds.series:
        return None
    i = bisect_left(xlsx_ds.dates, date)
    if i >= len(xlsx_ds.dates) or xlsx_ds.dates[i] != date:
        return None
    return xlsx_ds.series[tenor][i]


def merge_expected_close(
    sql_ds: Dataset | None,
    xlsx_ds: Dataset | None,
    expected: dt.date,
) -> Dataset:
    """기대 전영업일 하루에 대한 SQL·엑셀 병합. 순수 함수 — DB 도 파일도 안
    만진다. 테스트가 이 함수를 직접 친다.

    반환되는 데이터셋의 `source` 가 곧 판정이다 (Dataset.source 주석 참조).
    `data_key` 는 여기서 만들지 않는다 — 워터마크/파일 바이트는 I/O 라서
    `load_dataset_merged` 의 몫이다.
    """
    if sql_ds is None:
        if xlsx_ds is None:
            raise DataFileError(
                "IRS 종가를 어느 출처에서도 읽지 못했다 — MySQL 도, 엑셀도"
            )
        xlsx_ds.warnings.append(
            "SQL 을 읽지 못해 엑셀 전체로 폴백 — 1D 는 다른 계열이므로 짧은 끝"
            " 커브가 SQL 기준과 다르다"
        )
        return xlsx_ds  # source 는 이미 "xlsx"

    if sql_ds.asof >= expected:
        # SQL 이 기대일을 들고 있다(초과는 이론상 없지만 SQL 을 믿는다).
        if sql_ds.latest("1D") is not None:
            return sql_ds  # 정상 경로: 손대지 않는다
        patch = _xlsx_value_at(xlsx_ds, "1D", sql_ds.asof) if xlsx_ds else None
        if patch is not None:
            sql_ds.series["1D"][-1] = patch
            sql_ds.source = "sql+xlsx-1d"
            sql_ds.warnings.append(
                f"{sql_ds.asof} 의 1D 가 SQL 에 없어 엑셀 값 {patch} 로 채움"
                " [OWNER, 2026-08-11]"
            )
        else:
            sql_ds.warnings.append(
                f"{sql_ds.asof} 의 1D 가 SQL 에도 엑셀에도 없다 — 빈칸으로 서빙"
            )
        return sql_ds

    # SQL 에 기대일이 통째로 없다 → 그 하루를 엑셀에서 덧붙인다.
    if xlsx_ds is None or all(
        _xlsx_value_at(xlsx_ds, t, expected) is None for t in sql_ds.series
    ):
        sql_ds.warnings.append(
            f"기대 전영업일 {expected} 이 SQL 에도 엑셀에도 없다 — "
            f"{sql_ds.asof} 까지로 서빙 (지연 칩이 말한다)"
        )
        return sql_ds

    appended: dict[str, float | None] = {}
    absent: list[str] = []
    for tenor in sql_ds.series:
        v = _xlsx_value_at(xlsx_ds, tenor, expected)
        sql_ds.series[tenor].append(v)
        appended[tenor] = v
        if v is None:
            absent.append(tenor)
    sql_ds.dates.append(expected)
    sql_ds.source = "sql+xlsx-day"
    sql_ds.warnings.append(
        f"{expected} 종가가 SQL 에 없어 하루 전체를 엑셀에서 덧붙임"
        + (f" (엑셀에도 없는 노드: {', '.join(absent)})" if absent else "")
    )
    return sql_ds


def load_dataset_merged(
    xlsx_path: Path = DEFAULT_XLSX, today: dt.date | None = None
) -> Dataset:
    """서버와 정적 빌드가 함께 쓰는 **유일한 진입점** [OWNER, 2026-08-11].

    둘이 같은 로더를 지나야 static-agreement 게이트가 성립한다 — 한쪽만 병합을
    알면 폴백한 날마다 정적 트리와 라이브 API 가 갈라진다.

    `data_key` 는 여기서 채운다: 순수 SQL 이면 기존 워터마크 키 그대로(캐시
    연속성), 엑셀이 섞이면 병합분의 지문이 붙어 키가 갈라진다 — 내용이 다른데
    키가 같은 캐시가 이 프로젝트의 고질병이라서다.
    """
    import hashlib as _hashlib
    import json as _json

    today = today or market_today()
    expected = prev_kr_business_day(today)

    sql_ds: Dataset | None = None
    sql_err: Exception | None = None
    try:
        sql_ds = load_dataset_sql(today)
    except Exception as e:  # noqa: BLE001 — DB 다운도 폴백 사유다
        sql_err = e
        log.warning("[dataset] SQL 로드 실패, 엑셀 폴백 시도: %s", e)

    need_xlsx = (
        sql_ds is None
        or sql_ds.asof < expected
        or sql_ds.latest("1D") is None
    )
    xlsx_ds: Dataset | None = None
    if need_xlsx and Path(xlsx_path).exists():
        try:
            xlsx_ds = load_dataset(Path(xlsx_path), today)
        except DataFileError as e:
            log.warning("[dataset] 엑셀 보충 로드 실패: %s", e)

    if sql_ds is None and xlsx_ds is None:
        raise DataFileError(
            f"MySQL 도 엑셀도 읽지 못했다 (SQL: {sql_err})"
        ) from sql_err

    ds = merge_expected_close(sql_ds, xlsx_ds, expected)

    # 캐시 키. 순수 SQL = 기존 sql_data_hash 그대로 — 어제의 캐시가 오늘도
    # 맞는 한 살아 있어야 한다. 엑셀이 섞이면 병합분 값의 지문을 덧붙인다.
    from .cache import data_hash, sql_data_hash  # 지연 import (순환 없음)

    if ds.source == "xlsx":
        ds.data_key = data_hash(Path(xlsx_path), ds.asof)
    else:
        ds.data_key = sql_data_hash(ds.asof)
        if ds.source != "sql":
            merged = {t: ds.series[t][-1] for t in sorted(ds.series)}
            fp = _hashlib.sha256(
                _json.dumps(merged, sort_keys=True).encode()
            ).hexdigest()[:12]
            ds.data_key += f":{ds.source}:{fp}"
    return ds
